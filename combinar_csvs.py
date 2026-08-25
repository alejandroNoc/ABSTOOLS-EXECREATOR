"""
Combina varios archivos CSV (formato Timestamp / Valor / Tag) en una sola tabla,
alineados por Timestamp, rellenando los huecos con forward-fill y luego backward-fill.

PIEXTRACT - by Alejandro Burelo Sanchez

Requisitos:
    pip install pandas openpyxl pyexcelerate

Dos modos de uso:

  1) MODO INTERACTIVO (doble clic o "python combinar_csvs.py" sin argumentos):
     se abren diálogos para elegir la carpeta y dónde guardar el resultado
     (el diálogo deja elegir CSV o Excel).

  2) MODO SILENCIOSO / DESATENDIDO (para llamarlo desde VBA, un .exe, etc.):
         python combinar_csvs.py "C:\\ruta\\a\\la\\carpeta\\con\\csv" [motor]
     - No pregunta carpeta ni nombre de salida.
     - "motor" (opcional, default "csv"):
         "csv"           -> COMBINADO.csv (el mas rapido, sin limite de filas)
         "xlsx"          -> COMBINADO.xlsx usando pyexcelerate (mas rapido que
                             openpyxl; si no esta instalado, cae a openpyxl)
         "xlsx_openpyxl" -> COMBINADO.xlsx forzando openpyxl
     - Al terminar abre el archivo automáticamente (os.startfile, Windows).
     - Si algo falla, el error queda registrado en "COMBINADO_error.log"
       dentro de la misma carpeta.

En ambos modos se muestra una ventana con barra de progreso (verde) para
que se vea que el proceso está trabajando, ya que el .exe se compila con
--noconsole y de otra forma corre totalmente invisible.

Soporta dos formatos de CSV:
  1) Un solo tag por archivo: columnas Timestamp, Valor, Tag (en ese orden).
  2) Varios tags "en bloque" dentro del mismo archivo (como en el Excel
     original): grupos de 3 columnas (Timestamp, Valor, Tag) uno al lado
     del otro, separados por una columna vacía.
"""

import glob
import os
import re
import shutil
import sys
import tempfile
import time
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook

AUTOR = "PIEXTRACT BY: ALEJANDRO BURELO SANCHEZ"

# Excel (.xlsx) soporta como maximo 1,048,576 filas por hoja (incluyendo el
# encabezado). Dejamos un margen para el encabezado de cada hoja.
FILAS_MAX_POR_HOJA = 1_048_575

# Cada cuanto tiempo (segundos) se refresca la ventana mientras se escriben
# filas. No es un porcentaje fijo: el avance se ve libre/continuo, y esto
# solo evita redibujar la ventana mas seguido de lo que el ojo puede ver
# (lo cual ademas frenaria muchisimo la escritura con datasets grandes).
INTERVALO_REFRESCO_SEGUNDOS = 0.05


def guardar_excel_multi_hoja(df: pd.DataFrame, salida: str, progreso=None, paso_base: float = 0) -> int:
    """
    Guarda el DataFrame en un .xlsx escribiendo fila por fila con openpyxl
    (modo write_only, rapido incluso con millones de filas). Si supera el
    limite de una sola hoja de Excel, lo reparte automaticamente en
    "Datos1", "Datos2", etc. El progreso de cada hoja se reporta en la
    barra celeste de forma libre (no por saltos de 10%), refrescando la
    ventana varias veces por segundo mientras escribe. Devuelve la
    cantidad de hojas escritas.
    """
    total_filas_todas = len(df)

    if total_filas_todas <= FILAS_MAX_POR_HOJA:
        chunks = [df]
    else:
        n_hojas_calc = -(-total_filas_todas // FILAS_MAX_POR_HOJA)  # ceil
        chunks = [
            df.iloc[i * FILAS_MAX_POR_HOJA:(i + 1) * FILAS_MAX_POR_HOJA]
            for i in range(n_hojas_calc)
        ]

    n_hojas = len(chunks)
    wb = Workbook(write_only=True)
    filas_escritas_total = 0

    if progreso:
        progreso.mostrar_barra_hoja()

    for idx, chunk in enumerate(chunks, start=1):
        nombre_hoja = "Datos" if n_hojas == 1 else f"Datos{idx}"
        ws = wb.create_sheet(title=nombre_hoja)
        ws.append(list(chunk.columns))

        if progreso:
            valor_barra = paso_base + (filas_escritas_total / total_filas_todas)
            progreso.actualizar(valor_barra, f"Guardando hoja {idx}/{n_hojas}...")

        # NaN/NaT no son validos para openpyxl -> los pasamos a None (celda vacia)
        chunk_seguro = chunk.astype(object).where(pd.notnull(chunk), None)
        total_filas_hoja = len(chunk_seguro)
        ultimo_refresco = 0.0

        for i, fila in enumerate(chunk_seguro.itertuples(index=False, name=None), start=1):
            ws.append(fila)
            filas_escritas_total += 1

            if progreso:
                ahora = time.perf_counter()
                if ahora - ultimo_refresco >= INTERVALO_REFRESCO_SEGUNDOS or i == total_filas_hoja:
                    ultimo_refresco = ahora
                    valor_barra = paso_base + (filas_escritas_total / total_filas_todas)
                    progreso.actualizar(valor_barra)
                    progreso.actualizar_hoja(i, total_filas_hoja)

    if progreso:
        progreso.ocultar_barra_hoja()

    wb.save(salida)
    return n_hojas


def guardar_excel_pyexcelerate(df: pd.DataFrame, salida: str, progreso=None, paso_base: float = 0) -> int:
    """
    Guarda el DataFrame en un .xlsx usando la libreria pyexcelerate en vez
    de openpyxl. Es ~30-40% mas rapido que openpyxl para escribir, porque
    arma cada hoja de un solo golpe (bulk) en vez de fila por fila.
    Contrapartida: como escribe en bloque, el progreso dentro de cada hoja
    NO es granular fila-a-fila como con openpyxl -- se reporta por hoja
    completa (no hay barra celeste "Guardando Filas N/Total" para este
    motor). Igual que guardar_excel_multi_hoja, reparte en varias hojas
    ("Datos1", "Datos2", ...) si se pasa del limite de filas de Excel.
    """
    from pyexcelerate import Workbook as PyExcelerateWorkbook
    from pyexcelerate import Style, Format

    total_filas_todas = len(df)

    if total_filas_todas <= FILAS_MAX_POR_HOJA:
        chunks = [df]
    else:
        n_hojas_calc = -(-total_filas_todas // FILAS_MAX_POR_HOJA)  # ceil
        chunks = [
            df.iloc[i * FILAS_MAX_POR_HOJA:(i + 1) * FILAS_MAX_POR_HOJA]
            for i in range(n_hojas_calc)
        ]

    n_hojas = len(chunks)
    wb = PyExcelerateWorkbook()
    filas_escritas_total = 0

    for idx, chunk in enumerate(chunks, start=1):
        nombre_hoja = "Datos" if n_hojas == 1 else f"Datos{idx}"

        if progreso:
            valor_barra = paso_base + (filas_escritas_total / total_filas_todas)
            progreso.actualizar(
                valor_barra, f"Preparando datos para hoja {idx}/{n_hojas} (pyexcelerate)..."
            )

        # NaN/NaT no son validos -> los pasamos a None (celda vacia)
        chunk_seguro = chunk.astype(object).where(pd.notnull(chunk), None)
        datos = [list(chunk_seguro.columns)] + chunk_seguro.values.tolist()

        if progreso:
            valor_barra = paso_base + (filas_escritas_total / total_filas_todas)
            progreso.actualizar(
                valor_barra, f"Escribiendo hoja {idx}/{n_hojas} ({len(chunk_seguro):,} filas)..."
            )

        ws = wb.new_sheet(nombre_hoja, data=datos)
        # Columna 1 (Timestamp) con formato de fecha -- si no, Excel la
        # muestra como numero serial en vez de fecha legible.
        ws.set_col_style(1, Style(size=-1, format=Format("yyyy-mm-dd hh:mm:ss")))

        filas_escritas_total += len(chunk_seguro)
        if progreso:
            valor_barra = paso_base + (filas_escritas_total / total_filas_todas)
            progreso.actualizar(valor_barra, f"Hoja {idx}/{n_hojas} lista.")

    if progreso:
        progreso.actualizar(paso_base + 0.9, "Guardando archivo (pyexcelerate)...")

    wb.save(salida)
    return n_hojas


class VentanaProgreso:
    """Ventana simple con barra de progreso verde y el credito del autor
    arriba. Se actualiza llamando a .actualizar(valor, mensaje) y se cierra
    con .cerrar(). No usa hilos: hay que llamar a .actualizar(...) seguido
    para que la ventana se refresque (Tkinter sin mainloop propio)."""

    def __init__(self, total: int):
        self.total = max(total, 1)
        self.root = tk.Tk()
        self.root.title("PIEXTRACT")
        self.root.geometry("460x220")
        self.root.resizable(False, False)
        self.root.attributes("-topmost", True)

        tk.Label(
            self.root,
            text=AUTOR,
            font=("Segoe UI", 9, "bold"),
        ).pack(pady=(12, 6))

        self.label_estado = tk.Label(
            self.root, text="Iniciando...", font=("Segoe UI", 9)
        )
        self.label_estado.pack(pady=(0, 4))

        estilo = ttk.Style()
        try:
            estilo.theme_use("default")
        except Exception:
            pass
        estilo.configure(
            "Verde.Horizontal.TProgressbar",
            troughcolor="#e0e0e0",
            background="#2ecc71",
            bordercolor="#2ecc71",
            lightcolor="#2ecc71",
            darkcolor="#2ecc71",
            thickness=18,
        )
        estilo.configure(
            "Celeste.Horizontal.TProgressbar",
            troughcolor="#e0e0e0",
            background="#5dade2",
            bordercolor="#5dade2",
            lightcolor="#5dade2",
            darkcolor="#5dade2",
            thickness=14,
        )

        self.barra = ttk.Progressbar(
            self.root,
            style="Verde.Horizontal.TProgressbar",
            orient="horizontal",
            length=400,
            mode="determinate",
            maximum=self.total,
        )
        self.barra.pack(pady=(2, 2))

        self.label_pct = tk.Label(self.root, text="0%", font=("Segoe UI", 8))
        self.label_pct.pack(pady=(0, 8))

        # Barra secundaria (celeste), para el llenado fila a fila de cada
        # hoja del Excel. Se muestra/oculta con mostrar_barra_hoja() /
        # ocultar_barra_hoja() -- solo esta visible mientras se estan
        # escribiendo filas al archivo.
        self.frame_hoja = tk.Frame(self.root)
        self.label_hoja = tk.Label(self.frame_hoja, text="", font=("Segoe UI", 8))
        self.label_hoja.pack()
        self.barra_hoja = ttk.Progressbar(
            self.frame_hoja,
            style="Celeste.Horizontal.TProgressbar",
            orient="horizontal",
            length=400,
            mode="determinate",
            maximum=100,
        )
        self.barra_hoja.pack(pady=(4, 4))
        # frame_hoja no se empaqueta todavia -> arranca oculta

        self._refrescar()

    def actualizar(self, valor: float, mensaje: str = ""):
        valor = min(valor, self.total)
        self.barra["value"] = valor
        pct = int(round((valor / self.total) * 100))
        self.label_pct.config(text=f"{pct}%")
        if mensaje:
            self.label_estado.config(text=mensaje)
        self._refrescar()

    def mostrar_barra_hoja(self):
        self.frame_hoja.pack(pady=(0, 4))
        self._refrescar()

    def ocultar_barra_hoja(self):
        self.frame_hoja.pack_forget()
        self._refrescar()

    def actualizar_hoja(self, fila_actual: int, total_filas: int):
        pct = int(round((fila_actual / total_filas) * 100)) if total_filas else 0
        self.barra_hoja["value"] = pct
        self.label_hoja.config(
            text=f"Guardando Filas {fila_actual:,}/{total_filas:,} ({pct}%)"
        )
        self._refrescar()

    def _refrescar(self):
        try:
            self.root.update_idletasks()
            self.root.update()
        except Exception:
            pass

    def cerrar(self):
        try:
            self.root.destroy()
        except Exception:
            pass


def obtener_ruta_salida_disponible(carpeta: str, extension: str = ".csv") -> str:
    """
    Devuelve una ruta de salida lista para escribir dentro de 'carpeta'.
    Empieza probando "COMBINADO<extension>"; si ese archivo existe y esta
    BLOQUEADO (por ejemplo porque alguien lo tiene abierto en Excel),
    prueba "COMBINADO2<extension>", "COMBINADO3<extension>", etc. hasta
    encontrar uno que se pueda escribir. Si el archivo existe pero NO esta
    bloqueado, lo reusa igual (se sobrescribe, comportamiento de siempre).

    Default: .csv -- es ~5x mas rapido de escribir que .xlsx y no tiene
    limite de filas por hoja, asi que es el formato de salida por defecto.
    """
    intento = 0
    while True:
        nombre = f"COMBINADO{extension}" if intento == 0 else f"COMBINADO{intento + 1}{extension}"
        candidato = os.path.join(carpeta, nombre)

        if not os.path.exists(candidato):
            return candidato

        # El archivo ya existe: probamos si esta bloqueado intentando
        # abrirlo para escritura sin truncarlo (no modifica el contenido).
        try:
            with open(candidato, "r+b"):
                pass
            return candidato  # no estaba bloqueado -> lo sobrescribimos
        except (PermissionError, OSError):
            intento += 1
            if intento > 50:
                # Ultimo recurso para no quedar en loop infinito
                return os.path.join(carpeta, f"COMBINADO_{int(time.time())}{extension}")


def elegir_carpeta() -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    carpeta = filedialog.askdirectory(title="Selecciona la carpeta con los CSV")
    root.destroy()
    return carpeta


def elegir_archivo_salida(carpeta_inicial: str) -> str:
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    salida = filedialog.asksaveasfilename(
        title="Guardar archivo combinado como...",
        initialdir=carpeta_inicial,
        initialfile="COMBINADO.csv",
        defaultextension=".csv",
        filetypes=[("CSV (recomendado, mas rapido)", "*.csv"), ("Excel", "*.xlsx")],
    )
    root.destroy()
    return salida


def mostrar_error(mensaje: str, carpeta: str | None = None, modo_silencioso: bool = False):
    """Muestra el error en un messagebox (modo interactivo) o lo escribe
    en un log junto a la carpeta procesada (modo silencioso / exe)."""
    if modo_silencioso:
        try:
            log_path = os.path.join(carpeta or os.getcwd(), "COMBINADO_error.log")
            with open(log_path, "w", encoding="utf-8") as f:
                f.write(mensaje)
        except Exception:
            pass
    else:
        root = tk.Tk()
        root.withdraw()
        messagebox.showerror("Error", mensaje)
        root.destroy()


# ============================================================
# Parseo rapido de Timestamps.
# PI (via v.TimeStamp.LocalDate en el .bas) devuelve la fecha como texto
# con el formato del idioma/configuracion regional de Windows -- puede
# venir con nombre de mes en ingles ("25-Aug-26 10:00:00"), en espanol
# ("25-ago-26 10:00:00" o "25-ago.-26 10:00:00"), o en formato numerico
# ("25/08/2026 10:00:00"). Si no probamos un formato exacto, pandas cae a
# un parser fila-por-fila ~20x mas lento con datasets grandes. Por eso:
# 1) normalizamos nombres de mes (ingles/espanol) a numero de 2 digitos,
#    para no depender de que el locale del proceso coincida con el del
#    texto (evita fallas raras de referencia con %b de Python).
# 2) detectamos el formato exacto UNA SOLA VEZ (con una muestra chica) y
#    lo reusamos para todo el resto -- muchisimo mas rapido.
# ============================================================
_MESES_A_NUMERO = {
    "ene": "01", "jan": "01",
    "feb": "02",
    "mar": "03",
    "abr": "04", "apr": "04",
    "may": "05",
    "jun": "06",
    "jul": "07",
    "ago": "08", "aug": "08",
    "sep": "09", "sept": "09", "set": "09",
    "oct": "10",
    "nov": "11",
    "dic": "12", "dec": "12",
}

# El VBA (LanzarCombinadorEXE / SoloExtraer) ahora escribe el Timestamp
# directamente en ISO 8601, que es lo primero que probamos: pandas tiene
# un parser en C especifico para ISO, ~10-20x mas rapido que cualquier
# formato con nombre de mes en texto. El resto queda como respaldo por si
# se procesan CSVs viejos generados antes de este cambio.
_FORMATOS_FECHA_CANDIDATOS = [
    "%Y-%m-%d %H:%M:%S",
    "%d-%m-%y %H:%M:%S",
    "%d-%m-%Y %H:%M:%S",
    "%d/%m/%Y %H:%M:%S",
    "%d/%m/%y %H:%M:%S",
    "%m/%d/%Y %I:%M:%S %p",
    "%m/%d/%y %I:%M:%S %p",
]

# Cache global: una vez detectado el formato para esta corrida, se reusa
# para todos los archivos siguientes (todos vienen de la misma extraccion
# de PI, mismo locale).
_formato_fecha_cache: str | None = None


def _contiene_letras(valor: str) -> bool:
    return bool(re.search(r"[A-Za-z]", valor))


def _normalizar_nombres_de_mes_rapido(serie_texto: pd.Series) -> pd.Series:
    """
    Reemplaza el nombre del mes (ingles o espanol, con o sin punto) por su
    numero de 2 digitos. Usa slicing vectorizado + mapeo por diccionario
    (mucho mas rapido que una regex de alternancia sobre toda la columna).
    Si el primer valor no tiene letras (ya es numerico, p.ej. ISO 8601),
    no hace nada -- esto hace que los CSV nuevos (ISO) salten este paso
    por completo.
    """
    muestra = serie_texto.dropna()
    if len(muestra) == 0:
        return serie_texto

    primer_valor = str(muestra.iloc[0])
    if not _contiene_letras(primer_valor):
        return serie_texto  # ya es numerico/ISO, nada que normalizar

    m = re.search(r"[A-Za-z]{3,4}\.?", primer_valor)
    if not m:
        return serie_texto
    ini, fin = m.start(), m.end()

    serie_str = serie_texto.astype(str)
    prefijo = serie_str.str[:ini]
    mes_texto = serie_str.str[ini:fin].str.rstrip(".").str.lower()
    sufijo = serie_str.str[fin:]
    mes_numero = mes_texto.map(_MESES_A_NUMERO)
    mes_numero = mes_numero.fillna(mes_texto)  # si algo no matcheo, no romper
    return prefijo + mes_numero + sufijo


def _detectar_formato(muestra: pd.Series) -> str | None:
    muestra = muestra.dropna()
    if len(muestra) == 0:
        return None
    muestra = muestra.head(300)
    for fmt in _FORMATOS_FECHA_CANDIDATOS:
        try:
            parsed = pd.to_datetime(muestra, format=fmt, errors="coerce")
        except Exception:
            continue
        if parsed.notna().mean() >= 0.95:
            return fmt
    return None


def parsear_timestamps_rapido(serie_original: pd.Series) -> pd.Series:
    """Convierte una columna de texto de fechas a datetime, usando el
    formato detectado/cacheado cuando es posible (rapido) y cayendo al
    parser generico de pandas solo si hace falta (siempre correcto,
    aunque mas lento)."""
    global _formato_fecha_cache

    normalizada = _normalizar_nombres_de_mes_rapido(serie_original)

    if _formato_fecha_cache:
        parsed = pd.to_datetime(normalizada, format=_formato_fecha_cache, errors="coerce")
        # Si el formato cacheado deja de servir (archivo con otro formato),
        # volvemos a detectar en vez de devolver puros NaT.
        if len(normalizada.dropna()) == 0 or parsed.notna().mean() >= 0.90:
            return parsed

    fmt_detectado = _detectar_formato(normalizada)
    if fmt_detectado:
        _formato_fecha_cache = fmt_detectado
        return pd.to_datetime(normalizada, format=fmt_detectado, errors="coerce")

    # Fallback: parser generico de pandas (mas lento, pero siempre funciona)
    return pd.to_datetime(normalizada, errors="coerce")


def leer_csv_como_bloques(path: str) -> list[pd.DataFrame]:
    """
    Lee un CSV y devuelve una lista de DataFrames (Timestamp, <NombreTag>),
    uno por cada tag encontrado, ya sea que el archivo tenga un solo tag
    o varios en bloques de 3 columnas.
    """
    # dtype=str: ~2x mas rapido que dejar que pandas infiera el tipo de
    # cada columna. Como resultado "Valor" queda como texto y hay que
    # convertirlo a numerico a mano mas abajo (pd.to_numeric).
    raw = pd.read_csv(path, header=None, dtype=str)

    dfs = []
    col = 0
    while col + 2 < raw.shape[1]:
        bloque = raw.iloc[:, col:col + 3].copy()
        bloque.columns = ["Timestamp", "Valor", "Tag"]
        bloque = bloque.iloc[1:].dropna(how="all")  # saltar encabezado del bloque

        if len(bloque) > 0 and bloque["Tag"].notna().any():
            tag_name = bloque["Tag"].dropna().iloc[0]
            bloque = bloque[["Timestamp", "Valor"]].dropna()
            bloque["Valor"] = pd.to_numeric(bloque["Valor"], errors="coerce")
            bloque = bloque.dropna(subset=["Valor"])
            bloque.columns = ["Timestamp", tag_name]
            bloque["Timestamp"] = parsear_timestamps_rapido(bloque["Timestamp"])
            bloque = bloque.dropna(subset=["Timestamp"]).drop_duplicates(subset=["Timestamp"])
            if len(bloque) > 0:
                dfs.append(bloque)

        col += 4  # 3 columnas de datos + 1 de separación

    return dfs


def main():
    # --- Determinar modo: silencioso (con carpeta como argumento) o interactivo ---
    modo_silencioso = len(sys.argv) > 1
    carpeta = sys.argv[1] if modo_silencioso else elegir_carpeta()

    # --- Motor de guardado (solo aplica en modo silencioso; en modo
    # interactivo el usuario elige extension en el dialogo de guardar) ---
    # "csv"          -> pandas.to_csv (el mas rapido, sin limite de filas)
    # "xlsx"         -> pyexcelerate (si esta instalado) con fallback a
    #                    openpyxl si no esta disponible
    # "xlsx_openpyxl"-> fuerza openpyxl aunque pyexcelerate este instalado
    motor = sys.argv[2].strip().lower() if len(sys.argv) > 2 else "csv"
    if motor not in ("csv", "xlsx", "xlsx_openpyxl"):
        print(f"Motor '{motor}' no reconocido, se usa 'csv' por defecto.")
        motor = "csv"

    if not carpeta or not os.path.isdir(carpeta):
        msg = f"La carpeta no existe o no fue especificada: {carpeta!r}"
        print(msg)
        mostrar_error(msg, carpeta, modo_silencioso)
        os._exit(1)

    # Excluye el propio archivo de salida por si ya existe de una corrida anterior
    archivos_csv = sorted(
        f for f in glob.glob(os.path.join(carpeta, "*.csv"))
        if not os.path.basename(f).upper().startswith("COMBINADO")
    )
    if not archivos_csv:
        msg = "No se encontraron archivos .csv en esa carpeta."
        print(msg)
        mostrar_error(msg, carpeta, modo_silencioso)
        os._exit(1)

    print(f"Encontrados {len(archivos_csv)} archivos CSV:")
    for f in archivos_csv:
        print(f"  - {os.path.basename(f)}")

    # Total de pasos para la barra: 1 por archivo leido + 1 para combinar + 1 para guardar
    total_pasos = len(archivos_csv) + 2
    progreso = VentanaProgreso(total_pasos)

    todos_los_dfs = []
    for idx, f in enumerate(archivos_csv, start=1):
        nombre = os.path.basename(f)
        progreso.actualizar(idx - 1, f"Leyendo {nombre} ({idx}/{len(archivos_csv)})...")
        try:
            dfs = leer_csv_como_bloques(f)
            todos_los_dfs.extend(dfs)
            print(f"  {nombre}: {len(dfs)} tag(s) encontrados")
        except Exception as e:
            print(f"  Error leyendo {nombre}: {e}")
        progreso.actualizar(idx, f"Leido {nombre}")

    if not todos_los_dfs:
        progreso.cerrar()
        msg = "No se encontraron datos válidos (Timestamp/Valor/Tag) en los CSV."
        print(msg)
        mostrar_error(msg, carpeta, modo_silencioso)
        os._exit(1)

    progreso.actualizar(len(archivos_csv) + 1, "Combinando datos (merge por Timestamp)...")
    resultado = todos_los_dfs[0]
    for d in todos_los_dfs[1:]:
        resultado = resultado.merge(d, on="Timestamp", how="outer")

    resultado = resultado.sort_values("Timestamp").reset_index(drop=True)
    resultado = resultado.ffill().bfill()

    if modo_silencioso:
        extension_salida = ".csv" if motor == "csv" else ".xlsx"
        salida = obtener_ruta_salida_disponible(carpeta, extension=extension_salida)
        nombre_default = f"COMBINADO{extension_salida}"
        if os.path.basename(salida) != nombre_default:
            print(f"{nombre_default} estaba bloqueado/abierto, se usa: {os.path.basename(salida)}")
    else:
        progreso.cerrar()
        salida = elegir_archivo_salida(carpeta)
        if not salida:
            print("No se guardó ningún archivo.")
            os._exit(0)
        progreso = VentanaProgreso(total_pasos)

    progreso.actualizar(total_pasos - 1, "Guardando archivo combinado...")

    # Escribimos primero a un archivo temporal en el disco local (la carpeta
    # temporal de Windows, %TEMP%), NO en la carpeta final del usuario.
    # Motivo: si la carpeta final esta sincronizada con OneDrive (muy comun
    # que "Escritorio" este redirigido ahi sin que el usuario lo note),
    # cada escritura incremental dispara al cliente de sincronizacion y eso
    # frena mucho el proceso (se ve como que la barra "se traba" a mitad de
    # camino). Escribiendo local y moviendo el archivo ya terminado de un
    # solo golpe al final evitamos ese frenado.
    extension = ".xlsx" if salida.lower().endswith(".xlsx") else ".csv"
    fd_temp, ruta_temporal = tempfile.mkstemp(suffix=extension, prefix="piextract_")
    os.close(fd_temp)

    try:
        if salida.lower().endswith(".xlsx"):
            usar_pyexcelerate = motor == "xlsx"  # "xlsx_openpyxl" fuerza openpyxl
            n_hojas = None
            if usar_pyexcelerate:
                try:
                    n_hojas = guardar_excel_pyexcelerate(
                        resultado, ruta_temporal, progreso=progreso, paso_base=total_pasos - 1
                    )
                except ImportError:
                    print("pyexcelerate no esta instalado, se usa openpyxl como respaldo.")
                    usar_pyexcelerate = False
            if not usar_pyexcelerate:
                n_hojas = guardar_excel_multi_hoja(
                    resultado, ruta_temporal, progreso=progreso, paso_base=total_pasos - 1
                )
            if n_hojas and n_hojas > 1:
                print(
                    f"Datos repartidos en {n_hojas} hojas "
                    f"(limite de Excel: {FILAS_MAX_POR_HOJA} filas por hoja)."
                )
        else:
            resultado.to_csv(ruta_temporal, index=False)

        progreso.actualizar(total_pasos - 0.02, "Moviendo archivo a la carpeta final...")
        shutil.move(ruta_temporal, salida)
    finally:
        # Por si algo fallo despues de crear el temporal pero antes del move
        if os.path.exists(ruta_temporal):
            try:
                os.remove(ruta_temporal)
            except Exception:
                pass

    progreso.actualizar(total_pasos, "Listo.")
    print(f"Archivo combinado guardado en: {salida}")

    progreso.cerrar()

    # Abrir el archivo final automáticamente (Windows)
    try:
        os.startfile(salida)  # type: ignore[attr-defined]
    except Exception:
        pass  # si no es Windows o falla, simplemente no lo abre

    if not modo_silencioso:
        root = tk.Tk()
        root.withdraw()
        messagebox.showinfo("Listo", f"Archivo combinado guardado en:\n{salida}")
        root.destroy()

    # Cierre forzado del proceso.
    # Motivo: pandas/numpy (via las librerias de calculo que usan internamente,
    # p.ej. OpenBLAS) a veces dejan hilos secundarios vivos que no terminan
    # solos dentro de un .exe empaquetado con PyInstaller. Eso hace que el
    # proceso quede "colgado" en el Administrador de Tareas aunque ya termino
    # su trabajo (el archivo ya se genero y se abrio). os._exit() termina el
    # proceso de inmediato sin esperar a esos hilos ni correr el cleanup
    # normal del interprete -- es seguro aca porque ya no queda nada
    # pendiente por guardar.
    sys.stdout.flush() if sys.stdout else None
    os._exit(0)


if __name__ == "__main__":
    try:
        main()
    except Exception:
        error_completo = traceback.format_exc()
        print(error_completo)
        carpeta_fallback = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
        mostrar_error(error_completo, carpeta_fallback, modo_silencioso=len(sys.argv) > 1)
        os._exit(1)
